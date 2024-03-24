"""
Project: Automate boring stuffs in excel
- Create the excel file with template
- Fill the excel template with number (by excel functions)
- Format the excel
- Maintain and update the excel
- Move the excel to production
"""

import openpyxl
import subprocess
import metrics_dict
import job_metadata
import report_attributes

"""
List of Functions

"""


def find_cell_by_value(sheet_name, cell_value, col_index):
    quote_sheetname = openpyxl.utils.quote_sheetname(sheet_name.title)
    """Return the coordinate of a cell in a column based on its value"""
    for row in range(1, sheet_name.max_row + 1):
        if sheet_name.cell(row, col_index).value == cell_value:
            return f'={quote_sheetname}!{sheet_name.cell(row, col_index+1).coordinate}'
    else:
        raise ValueError(f'{cell_value} not found in {sheet_name.title}')


def fill_time_grid(sheet_name, time_cell, time_type, time_parameter):
    """
        Creating the week/month/quarter/year time series in the report
        time_cell: cell that contains run_date of the report
        time_type: generating series by week|month|quarter|year
        time_parameter: number of trailing week/month/quarter/year
    """
    start_point = max(10, sheet_name.max_column + 1)
    for index, column in enumerate(
            range(start_point, start_point + time_parameter)
            ):
        if time_type.lower() == 'week':
            sheet_name.cell(1, column).value = 7*(index - time_parameter)
            sheet_name.cell(2, column).value = \
                    f'={sheet_name.cell(1, column).coordinate}\
                    +{time_cell.coordinate}'
            sheet_name.cell(2, column).style = yyyymmdd

            sheet_name.cell(3, column).value = \
                    f'=_xlfn.ISOWEEKNUM({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(4, column).value = \
                    f'=_xlfn.MONTH({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(5, column).value = \
                    f'=_xlfn.ROUNDUP({sheet_name.cell(4, column).coordinate}/3,0)'

            sheet_name.cell(6, column).value = \
                    f'=YEAR({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(10, column).value = \
                    f'="Wk"&{sheet_name.cell(3, column).coordinate}\
                    &"-"&\
                    RIGHT({sheet_name.cell(6, column).coordinate},2)'
            format_cell(sheet_name.cell(10, column), bold=True, colour='FFFFFF', fill='000000', alignment='right')

        elif time_type.lower() == 'month':
            sheet_name.cell(1, column).value = 1*(index - time_parameter)

            sheet_name.cell(2, column).value = \
                    f'=_xlfn.EOMONTH({time_cell.coordinate},\
                    {sheet_name.cell(1, column).coordinate})'
            sheet_name.cell(2, column).style = yyyymmdd

            sheet_name.cell(4, column).value = \
                    f'=_xlfn.MONTH({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(5, column).value = f'=_xlfn.ROUNDUP({sheet_name.cell(4, column).coordinate}/3,0)'

            sheet_name.cell(6, column).value = f'=YEAR({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(10, column).value = \
                    f'=TEXT({sheet_name.cell(4, column).coordinate}*28, "mmm")\
                    &"-"&RIGHT({sheet_name.cell(6, column).coordinate},2)'
            format_cell(sheet_name.cell(10, column), bold=True, colour='FFFFFF', fill='000000', alignment='right')

        elif time_type.lower() == 'quarter':
            sheet_name.cell(1, column).value = 1*(index - time_parameter)

            sheet_name.cell(2, column).value = \
                    f'=_xlfn.EOMONTH({time_cell.coordinate},\
                    {sheet_name.cell(1, column).coordinate})'
            sheet_name.cell(2, column).style = yyyymmdd

            sheet_name.cell(4, column).value = \
                    f'=_xlfn.MONTH({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(5, column).value = f'=_xlfn.ROUNDUP({sheet_name.cell(4, column).coordinate}/3,0)'

            sheet_name.cell(6, column).value = f'=YEAR({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(10, column).value = \
                    f'=TEXT({sheet_name.cell(4, column).coordinate}*28, "mmm")\
                    &"-"&RIGHT({sheet_name.cell(6, column).coordinate},2)'
            format_cell(sheet_name.cell(10, column), bold=True, colour='FFFFFF', fill='000000', alignment='right')

        elif time_type.lower() == 'year':
            sheet_name.cell(1, column).value = 1*(index - time_parameter)

            sheet_name.cell(2, column).value = \
                    f'=_xlfn.EOMONTH({time_cell.coordinate},\
                    {sheet_name.cell(1, column).coordinate})'
            sheet_name.cell(2, column).style = yyyymmdd

            sheet_name.cell(4, column).value = \
                    f'=_xlfn.MONTH({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(5, column).value = f'=_xlfn.ROUNDUP({sheet_name.cell(4, column).coordinate}/3,0)'

            sheet_name.cell(6, column).value = f'=YEAR({sheet_name.cell(2, column).coordinate})'

            sheet_name.cell(10, column).value = \
                    f'=TEXT({sheet_name.cell(4, column).coordinate}*28, "mmm")\
                    &"-"&RIGHT({sheet_name.cell(6, column).coordinate},2)'
            format_cell(sheet_name.cell(10, column), bold=True, colour='FFFFFF', fill='000000', alignment='right')
    return None


def format_cell(cell, font_size=11, colour='000000', bold=False,
                fill=None, alignment='left', border=openpyxl.styles.Border()):
    cell.font = openpyxl.styles.Font(size=font_size, bold=bold, color=colour)
    cell.fill = (openpyxl.styles.PatternFill(fill_type=None) if fill is None
                 else openpyxl.styles.PatternFill(start_color=fill,
                                                  end_color=fill,
                                                  fill_type="solid")
                 )
    cell.alignment = openpyxl.styles.Alignment(horizontal=alignment)
    cell.border = border
    return None


""" * * * * * * * * * * * * * *

Format cell to in excel

* * * * * * * * * * * * * * * * """
mmm_yy = openpyxl.styles.NamedStyle(name='standardise_date',
                                    number_format='mmm-yy')

yyyymmdd = openpyxl.styles.NamedStyle(name='standardise_date',
                                    number_format='yyyy-mm-dd')

white_bold_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
bold_font = openpyxl.styles.Font(bold=True)
title_font = openpyxl.styles.Font(bold=True, size=20)

black_background_fill = openpyxl.styles.PatternFill(start_color="000000",
                                                    end_color="000000",
                                                    fill_type="solid")
center_aligned = openpyxl.styles.Alignment(horizontal="center")
indent_aligned = openpyxl.styles.Alignment(indent=1)

right_aligned = openpyxl.styles.Alignment(horizontal="right")

thick_border = openpyxl.styles.Side(border_style="thick")

square_border = openpyxl.styles.Border(top=thick_border,
                                       right=thick_border,
                                       left=thick_border,
                                       bottom=thick_border)


"""
Create the workbooks with required sheets to store template, raw and transformation
"""
wb = openpyxl.Workbook()
ws_main = wb['Sheet']
ws_main.title = 'Main'

wb.create_sheet('_METADATA_')
ws_metadata = wb['_METADATA_']
for row, (k, v) in enumerate(job_metadata.metadata.items(), start=1):
    ws_metadata.cell(row, 1).value = k
    ws_metadata.cell(row, 2).value = v



wb.create_sheet('Raw_One')
ws_raw = wb['Raw_One']

# Report Parameter
trailing_week = 5
trailing_month = 4
trailing_quarter = 2
trailing_year = 1


# Loading the metadata from the script

# set the position of metadata in Main Sheet
# A2 -> Time, A3 -> Country, A4 -> job_metadata
# B2 -> Date, B3 -> week, B4 -> month, B4 -> Quarter, B5 -> Year
# TODO: Make the reporting name dynamic


ws_main['$A$1'].value = 'REPORTING NAME HERE'
ws_main['$A$2'].value = find_cell_by_value(ws_metadata, 'RUN_DATE', 1)
ws_main['$B$2'].value = 'date'
ws_main['$A$3'].value = find_cell_by_value(ws_metadata, 'FREE_FORM', 1)
ws_main['$A$4'].value = find_cell_by_value(ws_metadata, 'JOB_ID', 1)
ws_main['$I3'].value = 'activity_week'
ws_main['$I4'].value = 'activity_month'
ws_main['$I5'].value = 'activity_quarter'
ws_main['$I6'].value = 'activity_year'
ws_main['$I$8'].value = 'REPORTING NAME HERE'
format_cell(ws_main['$I$8'], font_size=18, bold=True, alignment='left')


# Set up the time array
fill_time_grid(ws_main, ws_main['$A$2'], 'week', trailing_week)
fill_time_grid(ws_main, ws_main['$A$2'], 'month', trailing_month)
fill_time_grid(ws_main, ws_main['$A$2'], 'quarter', trailing_quarter)
fill_time_grid(ws_main, ws_main['$A$2'], 'year', trailing_year)


# Create Defined Names for metrics in SUMIFS

prefix_gms_metrics = []
for i in report_attributes.time_prefix:
    for metrics in report_attributes.gms_perf_metrics:
        metric_w_prefix = (str(i) + '_' + str(metrics))
        prefix_gms_metrics.append(metric_w_prefix)

ws_raw_columns = report_attributes.date_dimension + report_attributes.region_dimension + report_attributes.seller_origin_dimension + prefix_gms_metrics

for idx, c in enumerate(ws_raw_columns, start=1):
    ws_raw.cell(1, idx).value = c


sheet_title_quote = openpyxl.utils.quote_sheetname(ws_raw.title)
cell_range_list = []
for col in range(1, ws_raw.max_column+1):
    col_letter = openpyxl.utils.get_column_letter(col)
    anchor_cell_range = openpyxl.utils.absolute_coordinate(f'{col_letter}1:{col_letter}{ws_raw.max_row}')
    cell_range_list.append(f'{sheet_title_quote}!{anchor_cell_range}')


for idx, cell_range in zip(range(1, ws_raw.max_column+1), cell_range_list):
    name_defined = ws_raw.cell(1, idx).value
    defn = openpyxl.workbook.defined_name.DefinedName(name_defined, attr_text=cell_range)
    wb.defined_names[name_defined] = defn



# Row 11 -> Set up the metrics
# Col 1 -> Col 6 -> Set up the reporting dimension.
# Col 9 -> Print hierarchy of report


# Collect the metrics from the built-in metrics_dict
# TODO: make this metric dict more dynamic
# TODO: allow users to choose what metrics to include in the report
review_metrics = metrics_dict.gms + metrics_dict.units_sold

hierarchy = metrics_dict.metric_groupby_dimension(
        review_metrics,
        metrics_dict.arcs['Established'].keys()
)


starting_row = 11
for row, (met, regions) in enumerate(hierarchy.items()):
    for i, region in zip(
            range(starting_row+row*len(regions),
                  starting_row+(row+1)*len(regions) + row),
            regions):
        ws_main.cell(i, 1).value = 'wtd_' + met
        ws_main.cell(i, 2).value = region
        if region == 'AGG':
            ws_main.cell(i, 9).value = metrics_dict.metric_mapping[met]
            format_cell(ws_main.cell(i, 9), bold=True, fill='D3D3D3')
        else:
            ws_main.cell(i, 9).value = region
            format_cell(ws_main.cell(i, 9), alignment='center')


# Filling the formula

for idx, row in enumerate(ws_main.iter_rows(min_col=1, max_col=3, min_row=11), start=11):
    for col in range(10, 10+trailing_week+1):
        ws_main.cell(idx, col).value = f'=SUMIFS(\
                INDIRECT(${row[0].coordinate}),\
                activity_year,{ws_main.cell(6, col).coordinate},\
                activity_week,{ws_main.cell(3, col).coordinate},\
                region,${row[1].coordinate})'

        if ws_main.cell(idx, 9).value in metrics_dict.metric_mapping.values():
            format_cell(ws_main.cell(idx, col), bold=True, fill='D3D3D3', alignment='right')

    for col in range(10+trailing_week + 1, 10 + trailing_week + 1 + trailing_month + 1):
        ws_main.cell(idx, col).value = f'=SUMIFS(\
                INDIRECT(${row[0].coordinate}),\
                activity_year,{ws_main.cell(6, col).coordinate},\
                activity_month,{ws_main.cell(4, col).coordinate},\
                region,${row[1].coordinate})'
        if ws_main.cell(idx, 9).value in metrics_dict.metric_mapping.values():
            format_cell(ws_main.cell(idx, col), bold=True, fill='D3D3D3', alignment='right')


wb.save("main.xlsx")
subprocess.check_call(['open', '-a', 'Microsoft Excel', './main.xlsx'])
