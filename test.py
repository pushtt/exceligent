import openpyxl
import subprocess

wb = openpyxl.load_workbook('main.xlsx')

raw_one = wb['Raw_One']

openpyxl.utils.get_column_letter(raw_one.max_column)


sheet_title_quote = openpyxl.utils.quote_sheetname(raw_one.title)
cell_range_list = []
for col in range(1, raw_one.max_column+1):
    col_letter = openpyxl.utils.get_column_letter(col)
    anchor_cell_range = openpyxl.utils.absolute_coordinate(f'{col_letter}1:{col_letter}{raw_one.max_row}')
    cell_range_list.append(f'{sheet_title_quote}!{anchor_cell_range}')


for idx, cell_range in zip(range(1, raw_one.max_column+1), cell_range_list):
    name_defined = raw_one.cell(1, idx).value
    defn = openpyxl.workbook.defined_name.DefinedName(name_defined, attr_text=cell_range)
    wb.defined_names[name_defined] = defn

wb.save('main.xlsx')
subprocess.check_call(['open', '-a', 'Microsoft Excel', './main.xlsx'])
