import openpyxl
from datetime import datetime
import job_metadata
import subprocess

wb = openpyxl.Workbook()
mainwb = wb['Sheet']
mainwb.title = "DECK"
metadata = wb.create_sheet('__METADATA__')
for row, (k, v) in enumerate(job_metadata.metadata.items(), start=1):
    metadata.cell(row, 1).value = k
    metadata.cell(row, 2).value = v


class BusinessReview():
    def __init__(self):
        self.main = None
        self.metadata = None

    def lookup(self, lookup_col, lookup_value, return_col):
        ws = self.metadata
        col_index = 1
        """Return the coordinate of a cell in a column based on its value"""
        for row in range(1, ws.max_row + 1):
            try:
                if str(ws.cell(row, col_index).value).upper() == lookup_value.upper():
                    return f'={ws.title}!{ws.cell(row, return_col).coordinate}'
            except:
                raise ValueError(f'{lookup_value} not found in {ws.title}')

    def fill_info(self, date, seller_origin="ALL", region="ALL", marketplace="ALL", team="ALL"):
        ws = self.main
        ws.cell(1, 1).value = "SELLER ORIGIN"
        ws.cell(1, 2).value = seller_origin
        ws.cell(2, 1).value = "DATE"
        ws.cell(2, 2).value = date
        ws.cell(3, 1).value = "REGION"
        ws.cell(3, 2).value = region
        ws.cell(4, 1).value = "MARKETPLACE"
        ws.cell(4, 2).value = marketplace
        ws.cell(5, 1).value = "TEAM"
        ws.cell(5, 2).value = team
        return ws

    def generate_time_series(self, date: datetime.date, trailing: int, time_type: str, row_idx:int =2, col_idx:int =10):
        ws = self.main
        starting_point = max(col_idx, ws.max_column)
        for id, col in enumerate(range(starting_point, starting_point+trailing+1)):
            date_coor = ws.cell(2, 2).coordinate
            control_coor = ws.cell(1, col).coordinate

            if time_type.upper() == "WEEK":
                ws.cell(1, col).value = (id - trailing + 1)*7
                ws.cell(2, col).value = f"={date_coor} + {control_coor}"
                ws.cell(3, col).value = f"=_xlfn.ISOWEEKNUM({ws.cell(2, col).coordinate} + 1)"
            elif time_type.upper() == "MONTH":
                ws.cell(1, col).value = id - trailing
                ws.cell(2, col).value = f"=EOMONTH({date_coor}, {control_coor})"
                ws.cell(3, col).value = f"=MONTH({ws.cell(2, col).coordinate})"
            elif time_type.upper() == "QUARTER":
                ws.cell(1, col).value = id - trailing
                ws.cell(2, col).value = f"=EOMONTH({date_coor}, {control_coor})"
                ws.cell(3, col).value = f"=_xlfn.ROUNDUP(MONTH({ws.cell(2, col).coordinate})/3,0)"
            elif time_type.upper() == "YTD":
                ws.cell(1, col).value = id - trailing
                ws.cell(2, col).value = f"={date_coor} + {control_coor}"
                ws.cell(3, col).value = f"=_xlfn.ISOWEEKNUM({ws.cell(2, col).coordinate} + 1)"


deck = BusinessReview()
deck.metadata = metadata
deck.main = mainwb

seller_origin = deck.lookup(1, "FREE_FORM", 2)
date = deck.lookup(1, "RUN_DATE", 2)
deck.fill_info(date, seller_origin)
deck.generate_time_series(date, 5, "WEEK")
deck.generate_time_series(date, 4, "MONTH")
deck.generate_time_series(date, 2, "QUARTER")
deck.generate_time_series(date, 0, "YTD")

wb.save("main.xlsx")
subprocess.check_call(['open', '-a', 'Microsoft Excel', './main.xlsx'])
